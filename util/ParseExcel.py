import openpyxl
from openpyxl.styles import Border, Side, Font
import time
from config.VarConfig import *


class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def loadWorkBook(self, excelPathAndName):
        # 将Excel文件加载到内存，并获取其workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        # 根据sheet名获取sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        # 根据sheet的索引获取sheet对象
        try:
            sheetName = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetName)
        return sheet

    def getRowNumber(self, sheet):
        return sheet.max_row

    def getColsNumber(self, sheet):
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        return sheet.min_row

    def getStartColNumber(self, sheet):
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        try:
            return sheet.rows[rowNo-1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        # 获取sheet中某一列，返回的是这一列所有的数据内容组成tuple
        try:
            return sheet.columns[colNo-1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        # 根据单元格所在的位置索引获取该单元格中的值
        if coordinate !=None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def getCellOfObject(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        # 根据单元格所在位置索引获取该单元格对象
        if coordinate !=None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeOfCell(self, sheet, content, coordinate=None, rowNo=None, colsNo=None, style=None):
        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if style:
                    sheet.cell(row=rowNo, column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinate of cell!")


if __name__ == '__main__':
    pe = ParseExcel()
    pe.loadWorkBook(dataFilePath)
    sheet = pe.getSheetByIndex(1)
    print(sheet.title)
    print()













